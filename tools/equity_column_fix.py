#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
equity_column_fix.py - QmiaoAI 权益变动表列映射错误自动检测与修正工具 v1.0

问题描述:
  解析用友导出的权益变动表 Excel 时, 解析器可能因列索引偏移将
  "所有者权益合计"列的数据误读入"少数股东权益"键, 而将上年金额列
  (或空列)误读入"所有者权益合计"键。此问题在全资子公司报表中尤为
  明显(少数股东权益应为0, 若有非零值则说明列映射有误)。

  也兼容检测用友导出本身的列顺序错位问题。

检测原理:
  模式1 - check-json (推荐, 解析后检测):
    1. 读取解析后的 JSON 数据中的 equity_statement 部分
    2. 检查"所有者权益合计"键的值是否全为0
    3. 检查"少数股东权益"键是否有非零值
    4. 若同时满足, 判定为列映射错误
    5. 附加验证: 若提供资产负债表数据, 交叉验证"少数股东权益"列的
       期末值是否等于资产负债表所有者权益合计(确认数据被错放)

  模式2 - check (Excel 原始检测):
    1. 读取权益变动表 Excel sheet, 定位"少数股东权益"和"所有者权益合计"两列
    2. 检查"所有者权益合计"列是否全为0或null
    3. 检查"少数股东权益"列是否有非零值
    4. 若满足上述条件, 判定为列错位

  模式3 - fix (Excel 修正):
    交换 Excel 中错位的两列数据并保存

用法:
  python tools/equity_column_fix.py check-json <解析后JSON路径> [--balance-sheet <资产负债表JSON路径>]
  python tools/equity_column_fix.py check <excel文件路径> [--sheet <sheet名>]
  python tools/equity_column_fix.py fix <excel文件路径> --output <输出路径> [--sheet <sheet名>]

依赖:
  - openpyxl (Excel 读写, check/fix 模式需要)
  - Python 3.10+

退出码:
  0 = 无列映射问题
  1 = 检测到列映射错误 (check/check-json 模式) / 已修正 (fix 模式)
  2 = 文件或参数错误
"""
import sys
import os
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl 未安装, 请执行 pip install openpyxl", file=sys.stderr)
    sys.exit(2)

try:
    import json
except ImportError:
    sys.exit(2)

# ============================================================
# 列名匹配模式 - 兼容全角/半角/中文空格
# ============================================================
import unicodedata

def normalize_str(s):
    """全角转半角, 去除多余空格, 统一用于列名匹配"""
    if s is None:
        return ""
    s = str(s)
    # 全角转半角
    s = unicodedata.normalize("NFKC", s)
    # 去除所有空白
    s = "".join(s.split())
    return s.strip()

# 目标列名模式
PATTERN_MINORITY = ["少数股东权益", "少数股东权益合计"]
PATTERN_TOTAL_EQUITY = ["所有者权益合计", "所有者权益（或股东权益）合计", "股东权益合计"]
PATTERN_END_BALANCE = ["四、本期期末余额", "四、本年期末余额", "本年期末余额", "期末余额"]
PATTERN_BEGIN_BALANCE = ["二、本年期初余额", "上年期末余额", "期初余额"]


def match_column(header_value, patterns):
    """检查 header 值是否匹配任一模式(规范化后比较)"""
    norm = normalize_str(header_value)
    if not norm:
        return False
    for p in patterns:
        pn = normalize_str(p)
        if not pn:
            continue
        if pn in norm or norm in pn:
            return True
    return False


def find_equity_columns(ws):
    """
    在权益变动表中定位"少数股东权益"和"所有者权益合计"两列。

    返回: (minority_col, total_equity_col) 列号(1-based), 未找到返回 (None, None)
    """
    minority_col = None
    total_equity_col = None

    # 扫描前6行(用友报表通常有合并表头, 实际列名在第4-6行)
    for row_idx in range(1, min(7, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                continue
            if match_column(cell_val, PATTERN_MINORITY) and minority_col is None:
                minority_col = col_idx
            if match_column(cell_val, PATTERN_TOTAL_EQUITY) and total_equity_col is None:
                total_equity_col = col_idx

    return minority_col, total_equity_col


def find_data_rows(ws):
    """
    定位数据行(期末余额、期初余额、本期增减等关键行)。

    返回: dict {行标签: 行号(1-based)}
    """
    key_rows = {}
    for row_idx in range(1, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=2).value
        if first_cell is None:
            first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell is None:
            continue
        norm = normalize_str(first_cell)
        for pattern in PATTERN_END_BALANCE:
            if normalize_str(pattern) in norm:
                key_rows["期末余额"] = row_idx
                break
        for pattern in PATTERN_BEGIN_BALANCE:
            if normalize_str(pattern) in norm:
                key_rows["期初余额"] = row_idx
                break
        if "本期增减" in norm or "增减变动" in norm:
            key_rows["本期增减"] = row_idx
    return key_rows


def check_column_swap(ws, minority_col, total_equity_col):
    """
    检测是否存在列错位:
      - "所有者权益合计"列全为0或null
      - "少数股东权益"列有非零值

    返回: (is_swapped, details_dict)
    """
    if minority_col is None or total_equity_col is None:
        return False, {"error": "未找到目标列"}

    # 收集所有数据行(跳过表头行, 通常从第7行开始)
    data_start_row = 7
    minority_values = []
    total_equity_values = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        m_val = ws.cell(row=row_idx, column=minority_col).value
        t_val = ws.cell(row=row_idx, column=total_equity_col).value

        # 转为 float, None/文本视为0
        def to_float(v):
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(",", "").strip())
            except (ValueError, TypeError):
                return 0.0

        minority_values.append(to_float(m_val))
        total_equity_values.append(to_float(t_val))

    # 判断条件
    total_all_zero = all(v == 0.0 for v in total_equity_values)
    minority_has_nonzero = any(v != 0.0 for v in minority_values)

    is_swapped = total_all_zero and minority_has_nonzero

    # 期末余额行详情
    key_rows = find_data_rows(ws)
    end_row = key_rows.get("期末余额")
    end_minority = None
    end_total = None
    if end_row:
        def safe_float(row, col):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(",", "").strip()) if v else 0.0
            except (ValueError, TypeError):
                return 0.0
        end_minority = safe_float(end_row, minority_col)
        end_total = safe_float(end_row, total_equity_col)

    details = {
        "minority_col": minority_col,
        "total_equity_col": total_equity_col,
        "minority_col_letter": openpyxl.utils.get_column_letter(minority_col),
        "total_equity_col_letter": openpyxl.utils.get_column_letter(total_equity_col),
        "total_equity_all_zero": total_all_zero,
        "minority_has_nonzero": minority_has_nonzero,
        "end_balance_row": end_row,
        "end_minority_value": end_minority,
        "end_total_equity_value": end_total,
        "minority_nonzero_count": sum(1 for v in minority_values if v != 0.0),
        "total_nonzero_count": sum(1 for v in total_equity_values if v != 0.0),
    }

    return is_swapped, details


def swap_columns(ws, minority_col, total_equity_col, output_path):
    """
    交换两列的数据(仅数据行, 不动表头)。
    """
    wb = openpyxl.load_workbook(ws.parent.path, data_only=False)
    ws_out = wb[ws.title]

    data_start_row = 7
    for row_idx in range(data_start_row, ws_out.max_row + 1):
        m_cell = ws_out.cell(row=row_idx, column=minority_col)
        t_cell = ws_out.cell(row=row_idx, column=total_equity_col)
        m_val = m_cell.value
        t_val = t_cell.value
        m_cell.value = t_val
        t_cell.value = m_val

    wb.save(output_path)


def check_json(parsed_json_path, balance_json_path=None):
    """
    对已解析的 JSON 数据检测列错位。
    用于 AI 在对话中解析后调用验证。
    """
    with open(parsed_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if "equity_statement" not in data:
        return False, {"error": "JSON 中无 equity_statement 数据"}

    eq = data["equity_statement"]

    # 检查"所有者权益合计"列是否全为0
    total_equity_keys = ["所有者权益合计"]
    minority_keys = ["少数股东权益"]

    total_all_zero = True
    minority_has_nonzero = False

    for row_label, row_data in eq.items():
        if not isinstance(row_data, dict):
            continue
        for tk in total_equity_keys:
            tv = row_data.get(tk, 0)
            if isinstance(tv, (int, float)) and tv != 0:
                total_all_zero = False
        for mk in minority_keys:
            mv = row_data.get(mk, 0)
            if isinstance(mv, (int, float)) and mv != 0:
                minority_has_nonzero = True

    is_swapped = total_all_zero and minority_has_nonzero

    details = {
        "total_equity_all_zero": total_all_zero,
        "minority_has_nonzero": minority_has_nonzero,
    }

    # 如果提供了资产负债表, 做交叉验证
    if balance_json_path and is_swapped:
        with open(balance_json_path, "r", encoding="utf-8") as f:
            bs_data = json.load(f)
        bs_equity = None
        if isinstance(bs_data, dict):
            liab = bs_data.get("liabilities", bs_data.get("负债和所有者权益", {}))
            for k in ["所有者权益（或股东权益）合计", "所有者权益合计"]:
                if k in liab:
                    v = liab[k]
                    if isinstance(v, dict):
                        bs_equity = v.get("期末余额", v.get("期末", 0))
                    else:
                        bs_equity = v
                    break
        if bs_equity is not None:
            # 检查权益变动表"少数股东权益"列的期末值是否等于资产负债表所有者权益合计
            end_row = eq.get("四、本期期末余额", {})
            minority_end = end_row.get("少数股东权益", 0)
            if isinstance(minority_end, (int, float)) and isinstance(bs_equity, (int, float)):
                details["cross_check"] = {
                    "bs_total_equity": bs_equity,
                    "eq_minority_end": minority_end,
                    "match": abs(bs_equity - minority_end) < 0.02
                }

    return is_swapped, details


def main():
    parser = argparse.ArgumentParser(
        description="QmiaoAI 权益变动表列错位自动检测与修正工具"
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # check 命令 - 检查 Excel
    p_check = subparsers.add_parser("check", help="检测 Excel 权益变动表是否存在列错位")
    p_check.add_argument("file", help="Excel 文件路径")
    p_check.add_argument("--sheet", help="权益变动表 sheet 名(未指定时自动查找)")

    # fix 命令 - 修正 Excel
    p_fix = subparsers.add_parser("fix", help="修正 Excel 权益变动表列错位")
    p_fix.add_argument("file", help="Excel 文件路径")
    p_fix.add_argument("--output", required=True, help="修正后输出文件路径")
    p_fix.add_argument("--sheet", help="权益变动表 sheet 名(未指定时自动查找)")

    # check-json 命令 - 检查已解析的 JSON
    p_json = subparsers.add_parser("check-json", help="检测已解析 JSON 中的列错位")
    p_json.add_argument("file", help="解析后的 JSON 文件路径")
    p_json.add_argument("--balance-sheet", help="资产负债表 JSON 路径(用于交叉验证)")

    args = parser.parse_args()

    if args.command is None:
        parser.print_help()
        sys.exit(2)

    if args.command == "check":
        # 加载 Excel
        if not os.path.exists(args.file):
            print(f"ERROR: 文件不存在: {args.file}", file=sys.stderr)
            sys.exit(2)
        wb = openpyxl.load_workbook(args.file, data_only=True)

        # 定位权益变动表 sheet
        if args.sheet:
            if args.sheet not in wb.sheetnames:
                print(f"ERROR: sheet '{args.sheet}' 不存在", file=sys.stderr)
                sys.exit(2)
            ws = wb[args.sheet]
        else:
            ws = None
            for name in wb.sheetnames:
                norm = normalize_str(name)
                if "权益" in norm or "所有者" in norm:
                    ws = wb[name]
                    print(f"[Info] 自动定位 sheet: {name}")
                    break
            if ws is None:
                print("ERROR: 未找到权益变动表 sheet, 请用 --sheet 指定", file=sys.stderr)
                sys.exit(2)

        # 查找列
        minority_col, total_equity_col = find_equity_columns(ws)
        if minority_col is None or total_equity_col is None:
            print(f"[SKIP] 未找到'少数股东权益'或'所有者权益合计'列 (minority={minority_col}, total={total_equity_col})")
            sys.exit(0)

        # 检测
        is_swapped, details = check_column_swap(ws, minority_col, total_equity_col)

        if is_swapped:
            print("[ALERT] 检测到列错位!")
            print(f"  少数股东权益列: {details['minority_col_letter']}列 (col={details['minority_col']})")
            print(f"  所有者权益合计列: {details['total_equity_col_letter']}列 (col={details['total_equity_col']})")
            print(f"  所有者权益合计列全为0: {details['total_equity_all_zero']}")
            print(f"  少数股东权益列非零值数: {details['minority_nonzero_count']}")
            if details.get("end_balance_row"):
                print(f"  期末余额行(第{details['end_balance_row']}行):")
                print(f"    少数股东权益列值: {details['end_minority_value']:,.2f}")
                print(f"    所有者权益合计列值: {details['end_total_equity_value']:,.2f}")
            print("\n[建议] 使用 fix 命令修正, 或在解析时将两列对调读取")
            sys.exit(1)
        else:
            print("[PASS] 未检测到列错位")
            print(f"  少数股东权益列: {details.get('minority_col_letter', 'N/A')}")
            print(f"  所有者权益合计列: {details.get('total_equity_col_letter', 'N/A')}")
            print(f"  所有者权益合计列全为0: {details.get('total_equity_all_zero', 'N/A')}")
            print(f"  少数股东权益列非零值数: {details.get('minority_nonzero_count', 0)}")
            sys.exit(0)

    elif args.command == "fix":
        if not os.path.exists(args.file):
            print(f"ERROR: 文件不存在: {args.file}", file=sys.stderr)
            sys.exit(2)
        wb = openpyxl.load_workbook(args.file, data_only=True)

        if args.sheet:
            if args.sheet not in wb.sheetnames:
                print(f"ERROR: sheet '{args.sheet}' 不存在", file=sys.stderr)
                sys.exit(2)
            ws = wb[args.sheet]
        else:
            ws = None
            for name in wb.sheetnames:
                norm = normalize_str(name)
                if "权益" in norm or "所有者" in norm:
                    ws = wb[name]
                    break
            if ws is None:
                print("ERROR: 未找到权益变动表 sheet", file=sys.stderr)
                sys.exit(2)

        minority_col, total_equity_col = find_equity_columns(ws)
        if minority_col is None or total_equity_col is None:
            print("ERROR: 未找到目标列", file=sys.stderr)
            sys.exit(2)

        is_swapped, details = check_column_swap(ws, minority_col, total_equity_col)

        if not is_swapped:
            print("[PASS] 未检测到列错位, 无需修正")
            sys.exit(0)

        print("[ALERT] 检测到列错位, 开始修正...")
        swap_columns(ws, minority_col, total_equity_col, args.output)
        print(f"[DONE] 已修正并保存到: {args.output}")
        print(f"  交换了 {details['minority_col_letter']}列 <-> {details['total_equity_col_letter']}列 (数据行)")
        sys.exit(1)

    elif args.command == "check-json":
        if not os.path.exists(args.file):
            print(f"ERROR: 文件不存在: {args.file}", file=sys.stderr)
            sys.exit(2)

        bs_path = args.balance_sheet if args.balance_sheet else None
        is_swapped, details = check_json(args.file, bs_path)

        if is_swapped:
            print("[ALERT] JSON 数据中检测到列错位!")
            print(f"  所有者权益合计列全为0: {details['total_equity_all_zero']}")
            print(f"  少数股东权益列有非零值: {details['minority_has_nonzero']}")
            if "cross_check" in details:
                cc = details["cross_check"]
                print(f"  交叉验证: 资产负债表权益={cc['bs_total_equity']:,.2f}, 权益变动表少数股东权益列期末={cc['eq_minority_end']:,.2f}")
                print(f"  匹配: {'是' if cc['match'] else '否'}")
            print("\n[建议] 在解析时将'少数股东权益'和'所有者权益合计'两列对调读取")
            sys.exit(1)
        else:
            print("[PASS] JSON 数据未检测到列错位")
            sys.exit(0)


if __name__ == "__main__":
    main()
